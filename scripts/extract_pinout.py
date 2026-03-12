from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook


def main() -> None:
    workbook_path = Path("samples/Pinout.xlsx")
    output_path = Path("apps/web/public/samples/pinout.ladder_26pin.json")

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Ladder_26pin"]

    rows = []
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        p2, signal, awg, color, tp, p4 = row
        if p2 is None and p4 is None:
            continue
        rows.append(
            {
                "fromPin": int(p2) if p2 is not None else None,
                "toPin": int(p4) if p4 is not None else None,
                "signalName": signal or "",
                "leftLabel": signal or "",
                "rightLabel": signal or "",
                "awg": "" if awg is None else str(int(awg)),
                "color": color or "",
                "pair": tp or "",
                "type": "TP" if tp else "AWG",
                "used": bool(p2 and p4),
            }
        )

    output_path.write_text(json.dumps({"rows": rows, "pinCount": 26}, indent=2), encoding="utf-8")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
